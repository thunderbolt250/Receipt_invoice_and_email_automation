import argparse
import datetime as dt
import json
import os
import smtplib
from dataclasses import dataclass
from email.message import EmailMessage
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import simpleSplit
from reportlab.pdfgen import canvas

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass


REQUIRED_FIELDS = {
    "name": ["name", "full name", "student name", "student"],
    "email": ["email", "email address"],
    "amount": ["amount", "amount paid", "contribution", "paid"],
    "date": ["date", "payment date", "paid on", "received on"],
    "payment_method": ["payment method", "payment mode", "mode", "channel"],
}


@dataclass
class StudentPayment:
    name: str
    email: str
    amount: str
    date: str
    payment_method: str
    receipt_number: str


def normalize_header(value: str) -> str:
    return value.strip().lower()


def map_columns(headers):
    header_map = {normalize_header(h): h for h in headers if h}
    mapped = {}
    for field, aliases in REQUIRED_FIELDS.items():
        for alias in aliases:
            if alias in header_map:
                mapped[field] = header_map[alias]
                break
    missing = [field for field in REQUIRED_FIELDS if field not in mapped]
    if missing:
        raise ValueError(
            "Missing required columns: " + ", ".join(missing) + ". "
            "Update your Excel headers or adjust REQUIRED_FIELDS."
        )
    return mapped


def format_amount(value, currency):
    if value is None:
        return ""
    try:
        amount = float(value)
        return f"{amount:,.0f} {currency}"
    except (TypeError, ValueError):
        return f"{value} {currency}".strip()


def format_date(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%d %B %Y")
    if isinstance(value, dt.date):
        return value.strftime("%d %B %Y")
    return str(value)


def load_students(excel_path: Path, currency: str):
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    column_map = map_columns(headers)
    students = []
    for index, row in enumerate(rows[1:], start=2):
        row_data = dict(zip(headers, row))
        if not any(row_data.values()):
            continue
        name = row_data.get(column_map["name"], "").strip()
        email = row_data.get(column_map["email"], "").strip()
        amount_raw = row_data.get(column_map["amount"], "")
        date_raw = row_data.get(column_map["date"], "")
        payment_method = row_data.get(column_map["payment_method"], "") or "Mobile Money"
        if not name or not email:
            raise ValueError(f"Missing name or email in row {index}")
        receipt_number = f"R-{index:04d}"
        students.append(
            StudentPayment(
                name=name,
                email=email,
                amount=format_amount(amount_raw, currency),
                date=format_date(date_raw),
                payment_method=str(payment_method),
                receipt_number=receipt_number,
            )
        )
    return students


def load_json(path: Path):
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def draw_wrapped_text(canvas_obj, text, x, y, max_width, line_height):
    lines = simpleSplit(text, canvas_obj._fontname, canvas_obj._fontsize, max_width)
    for line in lines:
        canvas_obj.drawString(x, y, line)
        y -= line_height


def create_receipt_pdf(student: StudentPayment, template_path: Path, output_path: Path, config, positions):
    reader = PdfReader(str(template_path))
    template_page = reader.pages[0]
    page_width = float(template_page.mediabox.width)
    page_height = float(template_page.mediabox.height)

    overlay_path = output_path.with_suffix(".overlay.pdf")
    c = canvas.Canvas(str(overlay_path), pagesize=(page_width, page_height))
    c.setFont("Helvetica", 10)

    def place_field(key, value, line_height=12):
        if key not in positions["fields"]:
            return
        field = positions["fields"][key]
        max_width = field.get("max_width", page_width * 0.6)
        x = page_width * field["x_pct"]
        if positions.get("origin") == "top-left":
            y = page_height * (1 - field["y_pct"])
        else:
            y = page_height * field["y_pct"]
        draw_wrapped_text(c, value, x, y, max_width, line_height)

    place_field("receipt_number", student.receipt_number)
    place_field("date", student.date)
    place_field("received_from", student.name)
    place_field("amount", student.amount)
    place_field("amount_words", config.get("amount_words", ""))
    place_field("payment_method", student.payment_method)
    place_field("contribution_period", config.get("semester", ""))
    place_field("authorized_by", config.get("from_name", ""))

    c.save()

    overlay_reader = PdfReader(str(overlay_path))
    writer = PdfWriter()
    template_page.merge_page(overlay_reader.pages[0])
    writer.add_page(template_page)
    with output_path.open("wb") as handle:
        writer.write(handle)

    overlay_path.unlink(missing_ok=True)


def load_email_template(path: Path):
    with path.open("r", encoding="utf-8") as handle:
        return handle.read()


def build_email(student: StudentPayment, config, body_template, attachment_path: Path):
    msg = EmailMessage()
    msg["Subject"] = config.get("email_subject", "Official Receipt – Spring Semester Contribution")
    msg["From"] = f"{config.get('from_name')} <{config.get('from_email')}>"
    msg["To"] = student.email

    first_name = student.name.split()[0]
    body = body_template.format(
        student_first_name=first_name,
        student_name=student.name,
        amount=student.amount,
        payment_date=student.date,
        payment_method=student.payment_method,
        semester=config.get("semester"),
        from_name=config.get("from_name"),
        from_title=config.get("from_title"),
        from_email=config.get("from_email"),
    )
    msg.set_content(body)
    msg.add_attachment(
        attachment_path.read_bytes(),
        maintype="application",
        subtype="pdf",
        filename=attachment_path.name,
    )
    return msg


def send_email(message: EmailMessage, smtp_config):
    host = smtp_config["host"]
    port = smtp_config["port"]
    username = smtp_config["username"]
    password = smtp_config["password"]

    with smtplib.SMTP(host, port) as server:
        server.starttls()
        if username and password:
            server.login(username, password)
        server.send_message(message)


def load_smtp_config():
    return {
        "host": os.environ.get("SMTP_HOST", "smtp.gmail.com"),
        "port": int(os.environ.get("SMTP_PORT", "587")),
        "username": os.environ.get("SMTP_USER", ""),
        "password": os.environ.get("SMTP_PASSWORD", ""),
    }


def main():
    parser = argparse.ArgumentParser(description="Generate receipts and email them to students.")
    parser.add_argument("--excel", default="Student Funds List 2026-2027.xlsx")
    parser.add_argument("--template", default="Receipt Template.pdf")
    parser.add_argument("--output", default="receipts")
    parser.add_argument("--config", default="receipt_config.json")
    parser.add_argument("--positions", default="template_positions.json")
    parser.add_argument("--email-template", default="email_template.txt")
    parser.add_argument("--send", action="store_true", help="Actually send emails")
    args = parser.parse_args()

    config = load_json(Path(args.config))
    positions = load_json(Path(args.positions))
    body_template = load_email_template(Path(args.email_template))

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    students = load_students(Path(args.excel), config.get("currency", "RWF"))
    smtp_config = load_smtp_config()

    for student in students:
        output_path = output_dir / f"{student.name.replace(' ', '_')}_Receipt.pdf"
        create_receipt_pdf(student, Path(args.template), output_path, config, positions)
        message = build_email(student, config, body_template, output_path)
        if args.send:
            send_email(message, smtp_config)
        else:
            print(f"Prepared email to {student.email} with receipt {output_path}")


if __name__ == "__main__":
    main()
